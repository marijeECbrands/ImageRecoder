from __future__ import annotations

import base64
import hashlib
import io
import csv
import json
import os
import queue
import re
import shutil
import sys
import tempfile
import threading
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import customtkinter as ctk
from openai import OpenAI
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from tkinter import filedialog, messagebox, ttk

APP_NAME = "Fashion Image Renamer v6.2"
SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
DEFAULT_CATEGORIES = {
    1: "Model front",
    2: "Model back",
    3: "Model zijkant rechts",
    4: "Model zijkant links",
    5: "Detail artikel",
    6: "Detail stof",
    7: "Artikel front",
    8: "Artikel back",
}


@dataclass(frozen=True)
class ArticleInfo:
    supplier_code: str
    external_description: str
    article_number: str
    color_code: str
    row_number: int

    @property
    def group_key(self) -> str:
        return f"{self.row_number}:{self.article_number}:{self.color_code}"


@dataclass(frozen=True)
class MatchConfig:
    separator: str = "_"
    supplier_column: str = "Artikelnr. Leverancier"
    external_description_column: str = "Externe omschrijving"
    article_number_column: str = "Artikel nr."
    color_code_column: str = "Kleurcode"
    supplier_match: str = "exact"
    external_description_match: str = "contains"


@dataclass
class Result:
    source: Path
    supplier_code: str = ""
    external_description: str = ""
    article_number: str = ""
    color_code: str = ""
    category: int | None = None
    category_label: str = ""
    category_confidence: int = 0
    match_method: str = ""
    match_confidence: int = 0
    explanation: str = ""
    status: str = ""
    new_name: str = ""


def resource_path(filename: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / filename


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def find_column(headers: dict[int, str], accepted: set[str]) -> int | None:
    normalized = {normalize(v) for v in accepted}
    for col, header in headers.items():
        if normalize(header) in normalized:
            return col
    return None


def load_match_config(path: Path | None) -> MatchConfig:
    config_path = path if path and path.exists() else resource_path("matchconfig.json")
    if not config_path.exists():
        return MatchConfig()
    raw = json.loads(config_path.read_text(encoding="utf-8"))
    match = raw.get("match", raw) if isinstance(raw, dict) else {}
    columns = raw.get("columns", {}) if isinstance(raw, dict) else {}
    return MatchConfig(
        separator=str(match.get("filename_separator", "_") or "_"),
        supplier_column=str(columns.get("supplier", "Artikelnr. Leverancier")),
        external_description_column=str(columns.get("external_description", "Externe omschrijving")),
        article_number_column=str(columns.get("article_number", "Artikel nr.")),
        color_code_column=str(columns.get("color_code", "Kleurcode")),
        supplier_match=str(match.get("before_separator_match", "exact")).lower(),
        external_description_match=str(match.get("after_separator_match", "contains")).lower(),
    )


def load_mapping(excel_path: Path, config: MatchConfig) -> list[ArticleInfo]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    header_row = supplier_col = external_col = article_col = color_col = None
    for row_idx in range(1, min(ws.max_row, 15) + 1):
        headers = {c.column: clean_cell(c.value) for c in ws[row_idx]}
        supplier = find_column(headers, {config.supplier_column, "Artikelnr. Leverancier", "Artikelnummer leverancier", "Leveranciersartikelnummer"})
        external = find_column(headers, {config.external_description_column, "Externe omschrijving", "External description", "Omschrijving extern"})
        article = find_column(headers, {config.article_number_column, "Artikel nr.", "Artikel nr", "Artikelnummer", "Artikelnr.", "Nr."})
        color = find_column(headers, {config.color_code_column, "Kleurcode", "Kleur code", "Color code", "Colour code"})
        if supplier and external and article and color:
            header_row, supplier_col, external_col, article_col, color_col = (
                row_idx, supplier, external, article, color
            )
            break
    if not all((header_row, supplier_col, external_col, article_col, color_col)):
        raise ValueError(
            "Benodigde kolommen niet gevonden. Verwacht: 'Artikelnr. Leverancier', "
            "'Externe omschrijving', 'Artikel nr.' en 'Kleurcode'."
        )
    mapping: list[ArticleInfo] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        supplier = clean_cell(ws.cell(row_idx, supplier_col).value)
        external = clean_cell(ws.cell(row_idx, external_col).value)
        article = clean_cell(ws.cell(row_idx, article_col).value)
        color = clean_cell(ws.cell(row_idx, color_col).value)
        if supplier and external and article and color:
            mapping.append(ArticleInfo(supplier, external, article, color, row_idx))
    if not mapping:
        raise ValueError("De Excel bevat geen bruikbare regels.")
    return mapping


def split_filename_match_parts(path: Path, separator: str = "_") -> tuple[str, str] | None:
    """Haal alleen leverancier en externe code uit de bestandsnaam.

    Voorbeelden:
    - EX13-261070_EX153.jpg -> (EX13-261070, EX153)
    - EX11-261006_EX518_1.jpg -> (EX11-261006, EX518)

    Alles na het tweede scheidingsteken is een foto-index of aanvullende
    aanduiding en doet niet mee aan de Excel-match.
    """
    parts = [part.strip() for part in path.stem.split(separator)]
    if len(parts) < 2 or not parts[0] or not parts[1]:
        return None
    return parts[0], parts[1]


def external_code_variants(value: object) -> set[str]:
    """Maak equivalente varianten voor externe codes zoals 707 en EX707."""
    normalized = normalize(value)
    if not normalized:
        return set()

    variants = {normalized}
    if normalized.startswith("ex") and len(normalized) > 2:
        variants.add(normalized[2:])
    else:
        variants.add(f"ex{normalized}")
    return variants


def external_description_matches(code: object, description: object, mode: str) -> bool:
    """Vergelijk een externe code robuust met de externe omschrijving."""
    code_variants = external_code_variants(code)
    description_norm = normalize(description)
    if not code_variants or not description_norm:
        return False

    if mode == "exact":
        description_variants = external_code_variants(description)
        return bool(code_variants & description_variants)

    return any(variant in description_norm for variant in code_variants)


def build_supplier_index(rows: list[ArticleInfo]) -> dict[str, list[ArticleInfo]]:
    index: dict[str, list[ArticleInfo]] = {}
    for row in rows:
        index.setdefault(normalize(row.supplier_code), []).append(row)
    return index


def match_filename_to_excel(path: Path, rows: list[ArticleInfo], config: MatchConfig, supplier_index: dict[str, list[ArticleInfo]] | None = None) -> tuple[ArticleInfo | None, str, list[ArticleInfo]]:
    """Match vóór '_' exact op leverancier en na '_' als bevat-match in externe omschrijving."""
    parts = split_filename_match_parts(path, config.separator)
    if not parts:
        return None, f"Bestandsnaam bevat geen bruikbare '{config.separator}' voor de gecombineerde match.", []
    supplier_part, external_part = parts
    supplier_norm = normalize(supplier_part)
    if config.supplier_match == "contains":
        supplier_candidates = [row for row in rows if supplier_norm and supplier_norm in normalize(row.supplier_code)]
    else:
        if supplier_index is None:
            supplier_candidates = [row for row in rows if normalize(row.supplier_code) == supplier_norm]
        else:
            supplier_candidates = supplier_index.get(supplier_norm, [])

    matches = [
        row
        for row in supplier_candidates
        if external_description_matches(
            external_part, row.external_description, config.external_description_match
        )
    ]
    if len(matches) == 1:
        return matches[0], (
            f"Gecombineerde match: leverancier '{supplier_part}' exact en "
            f"externe omschrijving bevat '{external_part}'."
        ), matches
    if len(matches) > 1:
        return None, (
            f"Meerdere Excel-regels voldoen aan leverancier '{supplier_part}' en "
            f"externe omschrijving bevat '{external_part}'."
        ), matches
    if supplier_candidates:
        return None, (
            f"Leverancier '{supplier_part}' gevonden, maar '{external_part}' komt niet voor "
            "in Externe omschrijving."
        ), []
    return None, f"Leverancier '{supplier_part}' niet exact gevonden in Excel.", []

def load_categories(path: Path | None) -> dict[int, str]:
    if path and path.exists():
        raw = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(raw, dict) and "categories" in raw:
            raw = raw["categories"]
        categories = {int(k): str(v).strip() for k, v in raw.items() if str(v).strip()}
        if categories:
            return categories
    default = resource_path("fotocategorieen.json")
    if default.exists():
        raw = json.loads(default.read_text(encoding="utf-8"))
        if "categories" in raw:
            raw = raw["categories"]
        return {int(k): str(v) for k, v in raw.items()}
    return DEFAULT_CATEGORIES.copy()


def extract_images(source: Path, work: Path) -> list[Path]:
    if source.is_dir():
        return sorted(p for p in source.rglob("*") if p.suffix.lower() in SUPPORTED_EXTENSIONS)
    if source.suffix.lower() != ".zip":
        raise ValueError("Selecteer een ZIP-bestand of map met afbeeldingen.")
    target = work / "images"
    target.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(source) as archive:
        archive.extractall(target)
    return sorted(p for p in target.rglob("*") if p.suffix.lower() in SUPPORTED_EXTENSIONS)


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def perceptual_dhash(path: Path) -> int:
    """Snelle 64-bit dHash voor bijna-identieke foto's."""
    with Image.open(path) as image:
        image = image.convert("L").resize((9, 8))
        pixels = list(image.getdata())
    value = 0
    bit = 0
    for row in range(8):
        offset = row * 9
        for col in range(8):
            if pixels[offset + col] > pixels[offset + col + 1]:
                value |= 1 << bit
            bit += 1
    return value


def hamming_distance(a: int, b: int) -> int:
    return (a ^ b).bit_count()


def data_url(path: Path, max_px: int = 1200) -> str:
    """Maak een compacte AI-versie; het originele bestand wordt nooit aangepast."""
    with Image.open(path) as image:
        image = image.convert("RGB")
        if max(image.size) > max_px:
            image.thumbnail((max_px, max_px), Image.Resampling.LANCZOS)
        buffer = io.BytesIO()
        image.save(buffer, format="JPEG", quality=84, optimize=True)
    encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    return f"data:image/jpeg;base64,{encoded}"


def parse_json(text: str) -> dict:
    text = text.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    start, end = text.find("{"), text.rfind("}")
    if start < 0 or end < start:
        raise ValueError("Geen geldig JSON-object ontvangen van de AI.")
    return json.loads(text[start:end + 1])


def category_prompt(categories: dict[int, str]) -> str:
    return "; ".join(f"{number} {label}" for number, label in sorted(categories.items()))


def classify_group(
    client: OpenAI, model: str, files: list[Path], categories: dict[int, str], ai_max_px: int = 1200
) -> dict[str, dict]:
    content: list[dict] = [{
        "type": "input_text",
        "text": (
            "Beoordeel deze fashionfoto's als één artikelgroep en classificeer elk bestand. "
            f"Gebruik uitsluitend deze categorieën: {category_prompt(categories)}. "
            "Gebruik de foto's onderling als context. Categorie 5 (Detail artikel) is voor een specifiek onderdeel of constructiedetail "
            "zoals mouw, kraag, knopen, zak, zoom, rits of borduring. Categorie 6 (Detail stof) is voor een close-up waarbij "
            "materiaal, textuur, weving, print/patroon of het stofoppervlak centraal staat. Let extra goed op packshots: "
            "Artikel front is uitsluitend de voorkant van het kledingstuk zonder model; Artikel back is uitsluitend de achterkant "
            "zonder model. Een packshot van de achterkant mag nooit als Artikel front worden geclassificeerd. "
            "Kies bij twijfel de beste categorie met lagere confidence. "
            "Antwoord uitsluitend als JSON: "
            '{"results":[{"filename":"naam.jpg","category":1,"confidence":0,"explanation":"kort"}]}'
        ),
    }]
    for image in files:
        content.extend([
            {"type": "input_text", "text": f"Bestand: {image.name}"},
            {"type": "input_image", "image_url": data_url(image, ai_max_px)},
        ])
    response = client.responses.create(model=model, input=[{"role": "user", "content": content}])
    payload = parse_json(response.output_text)
    return {str(x.get("filename")): x for x in payload.get("results", [])}


def make_reference_sheet(groups: list[tuple[str, list[Path]]], target: Path, output: Path) -> dict[str, str]:
    thumb_w, thumb_h = 220, 260
    cols = 4
    rows = (len(groups) + cols - 1) // cols
    canvas = Image.new("RGB", (cols * thumb_w, rows * thumb_h + 70), "white")
    draw = ImageDraw.Draw(canvas)
    draw.text((10, 10), f"Doelfoto: {target.name}. Referentiegroepen hieronder.", fill="black")
    labels: dict[str, str] = {}
    for idx, (group_key, files) in enumerate(groups):
        token = f"R{idx + 1}"
        labels[token] = group_key
        x, y = (idx % cols) * thumb_w, 70 + (idx // cols) * thumb_h
        try:
            image = Image.open(files[0]).convert("RGB")
            image.thumbnail((thumb_w - 20, thumb_h - 55))
            canvas.paste(image, (x + 10, y + 30))
        except Exception:
            pass
        draw.rectangle((x, y, x + thumb_w - 1, y + thumb_h - 1), outline="black")
        draw.text((x + 8, y + 6), token, fill="black")
    canvas.save(output, quality=88)
    return labels


def visual_match(
    client: OpenAI,
    model: str,
    target: Path,
    known_groups: dict[str, list[Path]],
    work: Path,
    chunk_size: int = 12,
    ai_max_px: int = 1200,
) -> tuple[str | None, int, str]:
    candidates: list[tuple[str, int, str]] = []
    items = sorted(known_groups.items())
    for chunk_index in range(0, len(items), chunk_size):
        chunk = items[chunk_index:chunk_index + chunk_size]
        sheet = work / f"refs_{target.stem}_{chunk_index}.jpg"
        token_map = make_reference_sheet(chunk, target, sheet)
        content = [
            {
                "type": "input_text",
                "text": (
                    "Vergelijk de doelfoto met de referentiegroepen. Match alleen wanneer exact hetzelfde kledingartikel "
                    "én dezelfde kleur/uitvoering zichtbaar is. Een model, pose of achtergrond mag verschillen. "
                    "Gebruik NONE wanneer geen referentie overtuigend overeenkomt. Antwoord uitsluitend als JSON: "
                    '{"reference":"R1 of NONE","confidence":0-100,"explanation":"kort"}'
                ),
            },
            {"type": "input_text", "text": "Doelfoto:"},
            {"type": "input_image", "image_url": data_url(target, ai_max_px)},
            {"type": "input_text", "text": "Referentie-overzicht:"},
            {"type": "input_image", "image_url": data_url(sheet, ai_max_px)},
        ]
        response = client.responses.create(model=model, input=[{"role": "user", "content": content}])
        payload = parse_json(response.output_text)
        token = str(payload.get("reference", "NONE")).upper().strip()
        confidence = int(payload.get("confidence", 0) or 0)
        explanation = str(payload.get("explanation", ""))
        group_key = token_map.get(token)
        if group_key:
            candidates.append((group_key, confidence, explanation))
    if not candidates:
        return None, 0, "Geen visuele overeenkomst gevonden."
    return max(candidates, key=lambda x: x[1])


def classify_target_with_references(
    client: OpenAI,
    model: str,
    target: Path,
    references: list[Path],
    categories: dict[int, str],
    ai_max_px: int = 1200,
) -> tuple[int | None, int, str]:
    content: list[dict] = [{
        "type": "input_text",
        "text": (
            "Classificeer alleen de DOELFOTO. De overige foto's zijn referenties van hetzelfde artikel. "
            f"Gebruik uitsluitend: {category_prompt(categories)}. Antwoord uitsluitend als JSON: "
            '{"category":1,"confidence":0-100,"explanation":"kort"}'
        ),
    }, {"type": "input_text", "text": "DOELFOTO:"}, {"type": "input_image", "image_url": data_url(target, ai_max_px)}]
    for ref in references[:4]:
        content.extend([{"type": "input_text", "text": "Referentiefoto:"}, {"type": "input_image", "image_url": data_url(ref, ai_max_px)}])
    response = client.responses.create(model=model, input=[{"role": "user", "content": content}])
    payload = parse_json(response.output_text)
    try:
        category = int(payload.get("category"))
    except (TypeError, ValueError):
        category = None
    return category, int(payload.get("confidence", 0) or 0), str(payload.get("explanation", ""))



def adjudicate_same_category(
    client: OpenAI,
    model: str,
    first: Path,
    second: Path,
    category: int,
    categories: dict[int, str],
    ai_max_px: int,
) -> dict:
    """Extra controle wanneer twee beelden dezelfde categorie kregen.

    Voor Artikel front (7) kan de controle een beeld corrigeren naar Artikel back (8).
    Als beide werkelijk hetzelfde beeldtype/view zijn, wordt het tweede beeld als duplicaat beeldtype gemarkeerd.
    """
    label = categories.get(category, str(category))
    special = category == 7 and 8 in categories
    instruction = (
        "Twee foto's van hetzelfde fashionartikel kregen dezelfde categorie. Vergelijk uitsluitend deze twee beelden. "
        f"De huidige categorie is {category} ({label}). "
    )
    if special:
        instruction += (
            "Bepaal specifiek of één foto eigenlijk Artikel back is. "
            "Gebruik decision SECOND_IS_BACK als foto 2 de achterkant is en foto 1 de voorkant; "
            "FIRST_IS_BACK als foto 1 de achterkant is en foto 2 de voorkant; "
            "SAME_VIEW als beide in essentie hetzelfde beeldtype/dezelfde zijde tonen (bijvoorbeeld twee front-packshots); "
            "REVIEW als het niet betrouwbaar te bepalen is. "
        )
    else:
        instruction += (
            "Gebruik SAME_VIEW als beide in essentie hetzelfde beeldtype/dezelfde zijde tonen; "
            "REVIEW als ze inhoudelijk verschillende views zijn of als je niet betrouwbaar kunt beslissen. "
        )
    instruction += (
        'Antwoord uitsluitend als JSON: {"decision":"SAME_VIEW|SECOND_IS_BACK|FIRST_IS_BACK|REVIEW",'
        '"confidence":0-100,"explanation":"kort"}'
    )
    content = [
        {"type": "input_text", "text": instruction},
        {"type": "input_text", "text": f"Foto 1: {first.name}"},
        {"type": "input_image", "image_url": data_url(first, ai_max_px)},
        {"type": "input_text", "text": f"Foto 2: {second.name}"},
        {"type": "input_image", "image_url": data_url(second, ai_max_px)},
    ]
    response = client.responses.create(model=model, input=[{"role": "user", "content": content}])
    payload = parse_json(response.output_text)
    return {
        "decision": str(payload.get("decision", "REVIEW")).upper().strip(),
        "confidence": int(payload.get("confidence", 0) or 0),
        "explanation": str(payload.get("explanation", "")),
    }


def unique_destination(folder: Path, filename: str) -> Path:
    target = folder / filename
    if not target.exists():
        return target
    counter = 2
    while True:
        candidate = folder / f"{target.stem}_duplicaat{counter}{target.suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def safe_copy(result: Result, source: Path, folder: Path, filename: str) -> None:
    destination = unique_destination(folder, filename)
    shutil.copy2(source, destination)
    result.new_name = destination.name


def run_pipeline(
    image_source: Path,
    excel_path: Path,
    category_path: Path | None,
    match_config_path: Path | None,
    output_dir: Path,
    api_key: str,
    model: str,
    category_threshold: int,
    visual_match_threshold: int,
    progress: Callable[[int, int, str], None],
    max_workers: int = 4,
    ai_max_px: int = 1200,
    detect_duplicates: bool = True,
) -> tuple[list[Result], Path]:
    categories = load_categories(category_path)
    match_config = load_match_config(match_config_path)
    rows = load_mapping(excel_path, match_config)
    rows_by_key = {row.group_key: row for row in rows}
    supplier_index = build_supplier_index(rows)
    client = OpenAI(api_key=api_key)
    max_workers = max(1, min(int(max_workers), 8))
    ai_max_px = max(600, min(int(ai_max_px), 2000))

    if output_dir.exists():
        for name in ("hernoemd", "controle_nodig", "geen_match", "duplicaten", "artikel_back"):
            shutil.rmtree(output_dir / name, ignore_errors=True)
    good_dir = output_dir / "hernoemd"
    review_dir = output_dir / "controle_nodig"
    no_match_dir = output_dir / "geen_match"
    duplicate_dir = output_dir / "duplicaten"
    article_back_dir = output_dir / "artikel_back"
    for folder in (good_dir, review_dir, no_match_dir, duplicate_dir, article_back_dir):
        folder.mkdir(parents=True, exist_ok=True)

    cache_path = output_dir / "classification_cache.json"
    cache: dict[str, dict] = {}
    if cache_path.exists():
        try:
            loaded = json.loads(cache_path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                cache = loaded
        except Exception:
            cache = {}

    with tempfile.TemporaryDirectory(prefix="fashion_v6_") as temp_name:
        work = Path(temp_name)
        all_images = extract_images(image_source, work)
        if not all_images:
            raise ValueError("Geen ondersteunde afbeeldingen gevonden.")

        results: list[Result] = []

        # 0. Exacte duplicaten wereldwijd verwijderen vóór API-verwerking.
        images: list[Path] = []
        seen_exact: dict[str, Path] = {}
        if detect_duplicates:
            progress(0, 1, "Exacte duplicaten controleren…")
            for image in all_images:
                try:
                    digest = file_sha256(image)
                except Exception:
                    images.append(image)
                    continue
                original = seen_exact.get(digest)
                if original is None:
                    seen_exact[digest] = image
                    images.append(image)
                else:
                    result = Result(
                        source=image,
                        match_method="Duplicaatfilter",
                        match_confidence=100,
                        status="DUPLICAAT_EXACT",
                        explanation=f"Exact hetzelfde beeld als '{original.name}'. Niet naar AI gestuurd.",
                    )
                    safe_copy(result, image, duplicate_dir, image.name)
                    results.append(result)
        else:
            images = all_images

        matched_groups: dict[str, list[Path]] = {}
        direct_match_explanations: dict[Path, str] = {}
        unmatched: list[Path] = []
        ambiguous: dict[Path, str] = {}

        # 1. Zeer snelle Excel-match via index.
        for image in images:
            info, explanation, candidates = match_filename_to_excel(
                image, rows, match_config, supplier_index
            )
            if info:
                matched_groups.setdefault(info.group_key, []).append(image)
                direct_match_explanations[image] = explanation
            else:
                unmatched.append(image)
                if len(candidates) > 1:
                    ambiguous[image] = explanation

        # 2. Bijna-identieke foto's alleen binnen dezelfde betrouwbaar gematchte artikelgroep.
        if detect_duplicates:
            for group_key, files in list(matched_groups.items()):
                kept: list[Path] = []
                hashes: list[tuple[Path, int]] = []
                for image in files:
                    try:
                        dhash = perceptual_dhash(image)
                    except Exception:
                        kept.append(image)
                        continue
                    duplicate_of: Path | None = None
                    for ref, ref_hash in hashes:
                        if hamming_distance(dhash, ref_hash) <= 2:
                            duplicate_of = ref
                            break
                    if duplicate_of is None:
                        hashes.append((image, dhash))
                        kept.append(image)
                    else:
                        info = rows_by_key[group_key]
                        result = Result(
                            source=image,
                            supplier_code=info.supplier_code,
                            external_description=info.external_description,
                            article_number=info.article_number,
                            color_code=info.color_code,
                            match_method="Duplicaatfilter",
                            match_confidence=98,
                            status="DUPLICAAT_BIJNA_GELIJK",
                            explanation=(
                                f"Bijna identiek aan '{duplicate_of.name}' binnen hetzelfde artikel. "
                                "Niet naar AI gestuurd; origineel staat in map duplicaten."
                            ),
                        )
                        safe_copy(result, image, duplicate_dir, image.name)
                        results.append(result)
                matched_groups[group_key] = kept

        matched_groups = {k: v for k, v in matched_groups.items() if v}
        total = len(matched_groups) + len(unmatched)
        completed = 0
        used_names: set[str] = set()

        def classify_one_group(group_key: str, files: list[Path]):
            # Cache op beeldinhoud + model + categorie-indeling.
            classified: dict[str, dict] = {}
            uncached: list[Path] = []
            category_signature = json.dumps(categories, sort_keys=True, ensure_ascii=False)
            cache_keys: dict[Path, str] = {}
            for image in files:
                key = hashlib.sha256(
                    (file_sha256(image) + "|" + model + "|" + category_signature).encode("utf-8")
                ).hexdigest()
                cache_keys[image] = key
                item = cache.get(key)
                if isinstance(item, dict):
                    classified[image.name] = item
                else:
                    uncached.append(image)
            if uncached:
                fresh = classify_group(client, model, uncached, categories, ai_max_px)
                classified.update(fresh)
                for image in uncached:
                    item = fresh.get(image.name)
                    if isinstance(item, dict):
                        cache[cache_keys[image]] = item
            return group_key, files, classified

        # 3. Artikelgroepen parallel classificeren.
        group_outputs: dict[str, tuple[list[Path], dict[str, dict]] | Exception] = {}
        if matched_groups:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_map = {
                    executor.submit(classify_one_group, group_key, files): group_key
                    for group_key, files in sorted(matched_groups.items())
                }
                for future in as_completed(future_map):
                    group_key = future_map[future]
                    info = rows_by_key[group_key]
                    try:
                        _, files, classified = future.result()
                        group_outputs[group_key] = (files, classified)
                    except Exception as exc:
                        group_outputs[group_key] = exc
                    completed += 1
                    progress(
                        completed, total,
                        f"Classificatie klaar: {info.supplier_code} / {info.external_description}"
                    )

        # Resultaten in vaste volgorde samenstellen; originele foto's worden gekopieerd.
        for group_key, files in sorted(matched_groups.items()):
            info = rows_by_key[group_key]
            output = group_outputs.get(group_key)
            if isinstance(output, Exception) or output is None:
                exc = output if isinstance(output, Exception) else RuntimeError("Onbekende API-fout")
                for image in files:
                    result = Result(
                        source=image,
                        supplier_code=info.supplier_code,
                        external_description=info.external_description,
                        article_number=info.article_number,
                        color_code=info.color_code,
                        match_method="Bestandsnaam (2 velden)",
                        match_confidence=100,
                        status="API_FOUT",
                        explanation=str(exc),
                    )
                    safe_copy(result, image, review_dir, image.name)
                    results.append(result)
                continue

            _, classified = output

            # Eerst alle classificaties verzamelen. Bij een dubbele categorie volgt een extra AI-controle
            # voordat er bestanden worden gekopieerd of namen worden vastgelegd.
            prepared: list[dict] = []
            for image in files:
                item = classified.get(image.name, {})
                try:
                    category = int(item.get("category"))
                except (TypeError, ValueError):
                    category = None
                prepared.append({
                    "image": image,
                    "item": item,
                    "category": category,
                    "confidence": int(item.get("confidence", 0) or 0),
                    "forced_status": "",
                    "extra": "",
                })

            # Per categorie maximaal één normaal beeld. Een tweede 'Artikel front' wordt eerst expliciet
            # gecontroleerd op Artikel back. Is het werkelijk nogmaals front, dan is het DUPLICAAT_BEELDTYPE.
            first_for_category: dict[int, int] = {}
            for idx, entry in enumerate(prepared):
                category = entry["category"]
                if category not in categories:
                    continue
                # Detail artikel (5) en Detail stof (6) mogen meerdere legitieme beelden hebben.
                if category in (5, 6):
                    continue
                first_idx = first_for_category.get(category)
                if first_idx is None:
                    first_for_category[category] = idx
                    continue

                first_entry = prepared[first_idx]
                try:
                    verdict = adjudicate_same_category(
                        client, model, first_entry["image"], entry["image"],
                        category, categories, ai_max_px
                    )
                    decision = verdict["decision"]
                    vconf = verdict["confidence"]
                    vex = verdict["explanation"]
                    entry["extra"] = f"Dubbele-categoriecontrole {vconf}%: {vex}"

                    if category == 7 and 8 in categories and decision == "SECOND_IS_BACK":
                        entry["category"] = 8
                        entry["confidence"] = max(entry["confidence"], vconf)
                        if 8 in first_for_category:
                            entry["forced_status"] = "DUPLICAAT_BEELDTYPE"
                        else:
                            first_for_category[8] = idx
                    elif category == 7 and 8 in categories and decision == "FIRST_IS_BACK":
                        # De eerder geclassificeerde foto was de achterkant; corrigeer die naar categorie 8.
                        first_entry["category"] = 8
                        first_entry["confidence"] = max(first_entry["confidence"], vconf)
                        first_entry["extra"] = f"Herclassificatie naar Artikel back {vconf}%: {vex}"
                        first_for_category[7] = idx
                        if 8 in first_for_category and first_for_category[8] != first_idx:
                            first_entry["forced_status"] = "DUPLICAAT_BEELDTYPE"
                        else:
                            first_for_category[8] = first_idx
                    elif decision == "SAME_VIEW":
                        entry["forced_status"] = "DUPLICAAT_BEELDTYPE"
                    else:
                        entry["forced_status"] = "HERCLASSIFICATIE_NODIG"
                except Exception as exc:
                    entry["forced_status"] = "HERCLASSIFICATIE_NODIG"
                    entry["extra"] = f"Dubbele-categoriecontrole mislukt: {exc}"

            for entry in prepared:
                image = entry["image"]
                item = entry["item"]
                category = entry["category"]
                confidence = entry["confidence"]
                result = Result(
                    source=image,
                    supplier_code=info.supplier_code,
                    external_description=info.external_description,
                    article_number=info.article_number,
                    color_code=info.color_code,
                    category=category,
                    category_label=categories.get(category, ""),
                    category_confidence=confidence,
                    match_method="Bestandsnaam (2 velden)",
                    match_confidence=100,
                    explanation=(
                        f"{direct_match_explanations.get(image, '')} | "
                        f"Classificatie: {str(item.get('explanation', ''))} | "
                        f"{entry['extra']}"
                    ).strip(" |"),
                    status="OK",
                )
                if category not in categories:
                    result.status = "ONGELDIGE_CATEGORIE"
                    filename = image.name
                else:
                    filename = f"{info.article_number}-{info.color_code}-{category}{image.suffix.lower()}"
                    if confidence < category_threshold:
                        result.status = "CONTROLE_NODIG"
                    if entry["forced_status"]:
                        result.status = entry["forced_status"]

                # Een inhoudelijk duplicaat gaat naar duplicaten en krijgt niet kunstmatig _duplicaat2 in hernoemd.
                if result.status == "DUPLICAAT_BEELDTYPE":
                    safe_copy(result, image, duplicate_dir, image.name)
                else:
                    if filename.lower() in used_names:
                        result.status = "HERCLASSIFICATIE_NODIG"
                    used_names.add(filename.lower())
                    if result.status == "OK" and category == 8:
                        safe_copy(result, image, article_back_dir, filename)
                    else:
                        safe_copy(result, image, good_dir if result.status == "OK" else review_dir, filename)
                results.append(result)

        # 4. Visuele fallback parallel; alleen voor bestanden zonder betrouwbare naammatch.
        def visually_resolve(image: Path):
            initial_explanation = ambiguous.get(image, "Geen gecombineerde naammatch gevonden.")
            if not matched_groups:
                return image, None, 0, "", None, 0, "", initial_explanation
            group_key, match_conf, match_explanation = visual_match(
                client, model, image, matched_groups, work,
                ai_max_px=ai_max_px,
            )
            if not group_key or match_conf < visual_match_threshold:
                return image, group_key, match_conf, match_explanation, None, 0, "", initial_explanation
            category, cat_conf, cat_expl = classify_target_with_references(
                client, model, image, matched_groups[group_key], categories, ai_max_px
            )
            return image, group_key, match_conf, match_explanation, category, cat_conf, cat_expl, initial_explanation

        visual_outputs: dict[Path, tuple | Exception] = {}
        if unmatched:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_map = {executor.submit(visually_resolve, image): image for image in unmatched}
                for future in as_completed(future_map):
                    image = future_map[future]
                    try:
                        visual_outputs[image] = future.result()
                    except Exception as exc:
                        visual_outputs[image] = exc
                    completed += 1
                    progress(completed, total, f"Visuele fallback klaar: {image.name}")

        for image in unmatched:
            output = visual_outputs.get(image)
            initial_explanation = ambiguous.get(image, "Geen gecombineerde naammatch gevonden.")
            result = Result(source=image, match_method="Visuele vergelijking", explanation=initial_explanation)
            if not matched_groups:
                result.status = "GEEN_REFERENTIES"
                result.explanation += " Er zijn geen eenduidig op naam gematchte foto's om mee te vergelijken."
                safe_copy(result, image, no_match_dir, image.name)
                results.append(result)
                continue
            if isinstance(output, Exception) or output is None:
                result.status = "API_FOUT_VISUELE_MATCH"
                result.explanation = f"{initial_explanation} | {output}"
                safe_copy(result, image, review_dir, image.name)
                results.append(result)
                continue

            _, group_key, match_conf, match_explanation, category, cat_conf, cat_expl, initial_explanation = output
            result.match_confidence = match_conf
            result.explanation = f"{initial_explanation} | Visueel: {match_explanation}"
            if not group_key or match_conf < visual_match_threshold:
                result.status = "GEEN_BETROUWBARE_VISUELE_MATCH"
                safe_copy(result, image, no_match_dir, image.name)
            else:
                info = rows_by_key[group_key]
                result.supplier_code = info.supplier_code
                result.external_description = info.external_description
                result.article_number = info.article_number
                result.color_code = info.color_code
                result.category = category
                result.category_label = categories.get(category, "")
                result.category_confidence = cat_conf
                result.explanation += f" | Classificatie: {cat_expl}"
                result.status = "OK_VISUELE_MATCH"
                if category not in categories:
                    result.status = "ONGELDIGE_CATEGORIE"
                    filename = image.name
                else:
                    filename = f"{info.article_number}-{info.color_code}-{category}{image.suffix.lower()}"
                    if cat_conf < category_threshold:
                        result.status = "CONTROLE_NODIG"
                    if filename.lower() in used_names:
                        result.status = "DUBBELE_CATEGORIE_OF_NAAM"
                used_names.add(filename.lower())
                if result.status == "OK_VISUELE_MATCH" and category == 8:
                    safe_copy(result, image, article_back_dir, filename)
                else:
                    safe_copy(
                        result, image,
                        good_dir if result.status == "OK_VISUELE_MATCH" else review_dir,
                        filename,
                    )
            results.append(result)

    try:
        cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass

    report = output_dir / "hernoemrapport.csv"
    with report.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow([
            "Originele naam", "Artikelnr. Leverancier", "Externe omschrijving",
            "Artikel nr.", "Kleurcode", "Soort beeld nr.", "Soort beeld",
            "Classificatiezekerheid", "Matchmethode", "Matchzekerheid",
            "Status", "Nieuwe naam", "Toelichting",
        ])
        for r in results:
            writer.writerow([
                r.source.name, r.supplier_code, r.external_description,
                r.article_number, r.color_code, r.category or "", r.category_label,
                r.category_confidence, r.match_method, r.match_confidence,
                r.status, r.new_name, r.explanation,
            ])

    zip_path = output_dir / "resultaat_afbeeldingen.zip"
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for folder in (good_dir, review_dir, no_match_dir, duplicate_dir):
            for file in folder.rglob("*"):
                if file.is_file():
                    archive.write(file, file.relative_to(output_dir))
        archive.write(report, report.name)
    return results, zip_path


class App(ctk.CTk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1180x820")
        self.minsize(1000, 650)
        ctk.set_appearance_mode("system")
        self.image_source = ctk.StringVar()
        self.excel_path = ctk.StringVar()
        self.category_path = ctk.StringVar(value=str(resource_path("fotocategorieen.json")))
        self.match_config_path = ctk.StringVar(value=str(resource_path("matchconfig.json")))
        self.output_dir = ctk.StringVar(value=str(Path.home() / "FashionImageRenamer_Resultaat"))
        self.api_key = ctk.StringVar(value=os.environ.get("OPENAI_API_KEY", ""))
        self.model = ctk.StringVar(value="gpt-4.1-mini")
        self.category_threshold = ctk.IntVar(value=75)
        self.match_threshold = ctk.IntVar(value=85)
        self.max_workers = ctk.IntVar(value=4)
        self.ai_max_px = ctk.IntVar(value=1200)
        self.detect_duplicates = ctk.BooleanVar(value=True)
        self.events: queue.Queue = queue.Queue()
        self.all_results: list[Result] = []
        self.filter_text = ctk.StringVar()
        self.filter_status = ctk.StringVar(value="Alle statussen")
        self.is_processing = False
        self.processing_started_at = 0.0
        self.busy_frame_index = 0
        self._build_ui()
        self.filter_text.trace_add("write", lambda *_: self._apply_filters())
        self.after(100, self._poll)
        self.after(300, self._update_busy_indicator)

    def _row(self, parent, row, label, variable, command):
        ctk.CTkLabel(parent, text=label).grid(row=row, column=0, sticky="w", padx=14, pady=7)
        ctk.CTkEntry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=8, pady=7)
        ctk.CTkButton(parent, text="Kiezen", width=90, command=command).grid(row=row, column=2, padx=12, pady=7)

    def _build_ui(self):
        ctk.CTkLabel(self, text=APP_NAME, font=ctk.CTkFont(size=27, weight="bold")).pack(pady=(18, 8))
        ctk.CTkLabel(self, text="Snelle naammatch + parallelle AI + verkleinde AI-kopieën + duplicaatfilter | output blijft origineel").pack()
        form = ctk.CTkFrame(self)
        form.pack(fill="x", padx=22, pady=10)
        form.grid_columnconfigure(1, weight=1)
        self._row(form, 0, "Afbeeldingen (ZIP/map)", self.image_source, self._choose_images)
        self._row(form, 1, "Excel-sheet", self.excel_path, self._choose_excel)
        self._row(form, 2, "Fotocategorieën (JSON)", self.category_path, self._choose_categories)
        self._row(form, 3, "Matchconfig (JSON)", self.match_config_path, self._choose_match_config)
        self._row(form, 4, "Resultaatmap", self.output_dir, self._choose_output)
        ctk.CTkLabel(form, text="OpenAI API-key").grid(row=5, column=0, sticky="w", padx=14, pady=7)
        ctk.CTkEntry(form, textvariable=self.api_key, show="•").grid(row=5, column=1, sticky="ew", padx=8, pady=7)
        ctk.CTkLabel(form, text="Model").grid(row=6, column=0, sticky="w", padx=14, pady=7)
        ctk.CTkEntry(form, textvariable=self.model).grid(row=6, column=1, sticky="ew", padx=8, pady=7)
        limits = ctk.CTkFrame(form, fg_color="transparent")
        limits.grid(row=7, column=0, columnspan=3, sticky="ew", padx=14, pady=8)
        ctk.CTkLabel(limits, text="Classificatie ≥").pack(side="left")
        ctk.CTkEntry(limits, textvariable=self.category_threshold, width=55).pack(side="left", padx=(5, 20))
        ctk.CTkLabel(limits, text="Visuele match ≥").pack(side="left")
        ctk.CTkEntry(limits, textvariable=self.match_threshold, width=55).pack(side="left", padx=5)
        ctk.CTkLabel(limits, text="% (lager gaat naar controle/geen match)").pack(side="left")

        speed = ctk.CTkFrame(form, fg_color="transparent")
        speed.grid(row=8, column=0, columnspan=3, sticky="ew", padx=14, pady=(0, 8))
        ctk.CTkLabel(speed, text="Parallelle AI-taken").pack(side="left")
        ctk.CTkEntry(speed, textvariable=self.max_workers, width=50).pack(side="left", padx=(6, 18))
        ctk.CTkLabel(speed, text="AI max px").pack(side="left")
        ctk.CTkEntry(speed, textvariable=self.ai_max_px, width=65).pack(side="left", padx=(6, 18))
        ctk.CTkCheckBox(speed, text="Exacte + bijna-identieke beelden uitfilteren", variable=self.detect_duplicates).pack(side="left")

        buttons = ctk.CTkFrame(self, fg_color="transparent")
        buttons.pack(fill="x", padx=22, pady=6)
        self.start_btn = ctk.CTkButton(buttons, text="Start analyse", height=40, command=self._start)
        self.start_btn.pack(side="left")
        ctk.CTkButton(buttons, text="Open resultaatmap", height=40, command=self._open_output).pack(side="left", padx=10)
        self.progress = ctk.CTkProgressBar(self)
        self.progress.pack(fill="x", padx=22, pady=5)
        self.progress.set(0)

        activity = ctk.CTkFrame(self, fg_color="transparent")
        activity.pack(fill="x", padx=22, pady=(0, 5))
        self.busy_badge = ctk.CTkLabel(activity, text="● Gereed", font=ctk.CTkFont(weight="bold"))
        self.busy_badge.pack(side="left")
        self.status = ctk.CTkLabel(activity, text="Wacht op invoer")
        self.status.pack(side="left", padx=(12, 0))
        self.elapsed_label = ctk.CTkLabel(activity, text="")
        self.elapsed_label.pack(side="right")

        filter_bar = ctk.CTkFrame(self)
        filter_bar.pack(fill="x", padx=22, pady=(5, 4))
        ctk.CTkLabel(filter_bar, text="Filter resultaten").pack(side="left", padx=(12, 8), pady=8)
        self.filter_entry = ctk.CTkEntry(
            filter_bar, textvariable=self.filter_text,
            placeholder_text="Zoek op bestandsnaam, artikel, kleur, soort of nieuwe naam…", width=420
        )
        self.filter_entry.pack(side="left", fill="x", expand=True, pady=8)
        self.status_filter = ctk.CTkComboBox(
            filter_bar, variable=self.filter_status, values=["Alle statussen"],
            width=220, command=lambda _value: self._apply_filters()
        )
        self.status_filter.pack(side="left", padx=8, pady=8)
        ctk.CTkButton(filter_bar, text="Wis filter", width=90, command=self._clear_filters).pack(
            side="left", padx=(0, 8), pady=8
        )
        self.result_count = ctk.CTkLabel(filter_bar, text="0 resultaten", width=100)
        self.result_count.pack(side="right", padx=(0, 12))

        cols = ("origineel", "match", "leverancier", "extern", "artikel", "kleur", "soort", "zekerheid", "status", "nieuw")
        self.table = ttk.Treeview(self, columns=cols, show="headings", height=17)
        headings = ["Origineel", "Match", "Lev. artikel", "Externe omschrijving", "Artikel nr.", "Kleur", "Soort beeld", "Zekerheid", "Status", "Nieuwe naam"]
        widths = [190, 145, 125, 170, 85, 70, 140, 75, 170, 180]
        for col, heading, width in zip(cols, headings, widths):
            self.table.heading(col, text=heading)
            self.table.column(col, width=width)
        self.table.pack(fill="both", expand=True, padx=22, pady=(5, 18))

    def _choose_images(self):
        path = filedialog.askopenfilename(filetypes=[("ZIP", "*.zip"), ("Alle bestanden", "*.*")])
        if not path:
            path = filedialog.askdirectory(title="Kies map met afbeeldingen")
        if path: self.image_source.set(path)

    def _choose_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path: self.excel_path.set(path)

    def _choose_categories(self):
        path = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
        if path: self.category_path.set(path)

    def _choose_match_config(self):
        path = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
        if path: self.match_config_path.set(path)

    def _choose_output(self):
        path = filedialog.askdirectory()
        if path: self.output_dir.set(path)

    def _start(self):
        image_source, excel = Path(self.image_source.get()), Path(self.excel_path.get())
        output, categories = Path(self.output_dir.get()), Path(self.category_path.get())
        match_config = Path(self.match_config_path.get())
        if not image_source.exists() or not excel.exists():
            messagebox.showerror("Invoer ontbreekt", "Selecteer geldige afbeeldingen en Excel.")
            return
        if not self.api_key.get().strip():
            messagebox.showerror("API-key ontbreekt", "Vul de OpenAI API-key in.")
            return
        self.all_results = []
        self._clear_filters()
        for row in self.table.get_children(): self.table.delete(row)
        self.result_count.configure(text="0 resultaten")
        self.start_btn.configure(state="disabled")
        self.is_processing = True
        self.processing_started_at = time.monotonic()
        self.busy_frame_index = 0
        self.busy_badge.configure(text="● Bezig")
        self.status.configure(text="Bestanden voorbereiden…")
        self.elapsed_label.configure(text="Verstreken: 00:00")
        self.progress.set(0)
        threading.Thread(target=self._worker, args=(image_source, excel, categories, match_config, output), daemon=True).start()

    def _worker(self, images, excel, categories, match_config, output):
        try:
            results, zip_path = run_pipeline(
                images, excel, categories, match_config, output, self.api_key.get().strip(), self.model.get().strip(),
                int(self.category_threshold.get()), int(self.match_threshold.get()),
                lambda done, total, text: self.events.put(("progress", done, total, text)),
                max_workers=int(self.max_workers.get()),
                ai_max_px=int(self.ai_max_px.get()),
                detect_duplicates=bool(self.detect_duplicates.get()),
            )
            self.events.put(("done", results, zip_path))
        except Exception as exc:
            self.events.put(("error", str(exc)))

    def _poll(self):
        try:
            while True:
                event = self.events.get_nowait()
                if event[0] == "progress":
                    _, done, total, text = event
                    self.progress.set(done / max(total, 1)); self.status.configure(text=text)
                elif event[0] == "done":
                    _, results, zip_path = event
                    self.all_results = results
                    statuses = sorted({r.status for r in results if r.status})
                    self.status_filter.configure(values=["Alle statussen", *statuses])
                    self.filter_status.set("Alle statussen")
                    self._apply_filters()
                    self.is_processing = False
                    self.progress.set(1)
                    self.busy_badge.configure(text="● Gereed")
                    self.status.configure(text=f"Klaar: {len(results)} afbeeldingen verwerkt")
                    self.start_btn.configure(state="normal")
                    messagebox.showinfo("Klaar", f"Resultaat gemaakt:\n{zip_path}")
                elif event[0] == "error":
                    self.is_processing = False
                    self.start_btn.configure(state="normal")
                    self.busy_badge.configure(text="● Fout")
                    self.status.configure(text="Verwerking gestopt")
                    messagebox.showerror("Fout", event[1])
        except queue.Empty:
            pass
        self.after(100, self._poll)

    def _result_values(self, r: Result) -> tuple[str, ...]:
        return (
            r.source.name, f"{r.match_method} {r.match_confidence}%", r.supplier_code,
            r.external_description, r.article_number, r.color_code, f"{r.category or ''} {r.category_label}".strip(),
            f"{r.category_confidence}%", r.status, r.new_name,
        )

    def _apply_filters(self) -> None:
        if not hasattr(self, "table"):
            return
        query = normalize(self.filter_text.get())
        selected_status = self.filter_status.get()
        for row in self.table.get_children():
            self.table.delete(row)
        visible = 0
        for result in self.all_results:
            values = self._result_values(result)
            searchable = normalize(" ".join(values))
            if query and query not in searchable:
                continue
            if selected_status != "Alle statussen" and result.status != selected_status:
                continue
            self.table.insert("", "end", values=values)
            visible += 1
        total = len(self.all_results)
        self.result_count.configure(text=f"{visible} van {total}")

    def _clear_filters(self) -> None:
        self.filter_text.set("")
        self.filter_status.set("Alle statussen")
        self._apply_filters()

    def _update_busy_indicator(self) -> None:
        if self.is_processing:
            frames = ("● Bezig", "● Bezig.", "● Bezig..", "● Bezig...")
            self.busy_badge.configure(text=frames[self.busy_frame_index % len(frames)])
            self.busy_frame_index += 1
            elapsed = max(0, int(time.monotonic() - self.processing_started_at))
            minutes, seconds = divmod(elapsed, 60)
            self.elapsed_label.configure(text=f"Verstreken: {minutes:02d}:{seconds:02d}")
        self.after(500, self._update_busy_indicator)

    def _open_output(self):
        path = Path(self.output_dir.get()); path.mkdir(parents=True, exist_ok=True)
        os.startfile(path)  # type: ignore[attr-defined]


if __name__ == "__main__":
    App().mainloop()
